from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd 
import glob
from datetime import datetime

f = open('../Logs/Logger.txt',"a+")
f.write('The code is run at Main Script '+datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ '\n')


def avg_frequency(lst):
    return (sum(lst)/len(lst))

def band_duration_info():
    below_band = 0
    in_the_band = 0
    above_band = 0
    percent_value = 86.40
    max_frequency = round(max(frequency_value),2)

    min_frequency = round(min(frequency_value),2)

    avg_freq = round(avg_frequency(frequency_value),2)

    for each in frequency_value:
        if each < 49.90:
            below_band += 1
        elif each >= 49.90 and each <= 50.05:
            in_the_band += 1
        else:
            above_band += 1
    
    outside_band = round(below_band/percent_value + above_band/percent_value ,2)
    hours_outside_band = round( (outside_band/100) * 24 ,2)
    band_duration_info = {
        'max_freq':max_frequency,
        'min_freq':min_frequency,
        'avg_freq':avg_freq,
        'below_band':round(below_band/percent_value,2),
        'in_the_band':round(in_the_band/percent_value,2),
        'above_band':round(above_band/percent_value,2),
        'outside_band':outside_band,
        'hours_outside_band':hours_outside_band,
        'fdi': round(hours_outside_band/24,2),
        'date':date,
        
    }
    return band_duration_info

def weekly_fdi():
    totalHoursOutOfIEGC = 0
    for eachFrequencyProfile in frequency_profile_data:
          totalHoursOutOfIEGC += eachFrequencyProfile['hours_outside_band']
    return round(totalHoursOutOfIEGC/(168) , 2)

files = glob.glob("../Files/*.xlsx")
frequency_profile_data = []

for file in files:
    wb = load_workbook(filename = file)
    ws = wb.active
    frequency_cell = ws['B2:B8641']
    frequency_value = []
    dateRow = ws["A1"]
    date = datetime.strftime(dateRow.value,'%d')
    
    for row in frequency_cell:
        for col in row:
             frequency_value.append(col.value)
             

    frequency_profile_data.append(band_duration_info())












