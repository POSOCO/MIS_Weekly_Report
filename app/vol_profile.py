from openpyxl import Workbook
from openpyxl import load_workbook
import glob
from datetime import datetime


f = open('../Logs/Logger.txt',"a+")
f.write('The code is run at Vol Profile Script '+datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ '\n')

files = glob.glob("../Files/vol/*.xlsx")

vol_station_list_400KV = {'Bhopal':'P','Khandwa':'G','Itarsi':'F','Damoh':'AZ','Nagda':'O','Indore':'N','Gwalior':'AS','Raipur':'BL','Raigarh':'BK','Bhilai':'AQ','Wardha':'AV','Dhule':'AE','Parli':'AC','Boisar':'CA','Kalwa':'V','Karad':'AA','Asoj':'AH','Dehgam':'L','Kasor':'AO',
'Jetpur':'AJ','Amreli':'BQ','Vapi':'BZ'}
vol_station_list_765KV = {'Sipat':'AX','Seoni':'AW','Wardh':'BD','Bina':'BC','Indore':'BF','Sasan':'BH','Satna':'BG','Tamnar':'BW','Kotra':'BX','Vododara':'CF','Durg':'BY','Gwalior':'BB'}


vol_station_list = {'Bhopal':'P','Khandwa':'G','Itarsi':'F','Damoh':'AZ','Nagda':'O','Indore':'N','Gwalior':'AS','Raipur':'BL','Raigarh':'BK','Bhilai':'AQ','Wardha':'AV','Dhule':'AE','Parli':'AC','Boisar':'CA','Kalwa':'V','Karad':'AA','Asoj':'AH','Dehgam':'L','Kasor':'AO',
'Jetpur':'AJ','Amreli':'BQ','Vapi':'BZ','Sipat':'AX','Seoni':'AW','Wardh':'BD','Bina':'BC','Ind765':'BF','Sasan':'BH','Satna':'BG','Tamnar':'BW','Kotra':'BX','Vododara':'CF','Durg':'BY','Gwa765':'BB'}



vol_profile_data_400KV = []
vol_profile_data_765KV = []
vol_profile_data = {}
dateList = []

for eachStation in vol_station_list:
    vol_profile_data[eachStation] = {}

for file in files:
    wb = load_workbook(filename = file)
    ws = wb.active
    dateRow = ws["A4"]
    date = dateRow.value.split('-', maxsplit=1)[0]
    dateList.append(date)
    # for eachStation in vol_station_list_400KV:
    #     cell_location_min = vol_station_list_400KV[eachStation] + '1444'
    #     vol_cell_min = ws[cell_location_min]

    #     cell_location_max = vol_station_list_400KV[eachStation] + '1446'
    #     vol_cell_max = ws[cell_location_max]
        
    #     obj = {}
    #     obj[eachStation] = {
    #         'Min':vol_cell_min.value,
    #         'Max':vol_cell_max.value
    #     }
        
    #     vol_profile_data_400KV.append(obj)
    
    # for eachStation in vol_station_list_765KV:
    #     cell_location_min = vol_station_list_765KV[eachStation] + '1444'
    #     vol_cell_min = ws[cell_location_min]

    #     cell_location_max = vol_station_list_765KV[eachStation] + '1446'
    #     vol_cell_max = ws[cell_location_max]

    #     obj = {}
    #     obj[eachStation] = {
    #         'Min':vol_cell_min.value,
    #         'Max':vol_cell_max.value
    #     }
        
    #    vol_profile_data_765KV.append(obj)

    for eachStation in vol_station_list:
       
        cell_location_min = vol_station_list[eachStation] + '1444'
        vol_cell_min = ws[cell_location_min]
        cell_location_max = vol_station_list[eachStation] + '1446'
        vol_cell_max = ws[cell_location_max]
        
        
        vol_profile_data[eachStation].update({
            date :{
                 'Min':vol_cell_min.value,
                'Max':vol_cell_max.value
            }
        })

    

